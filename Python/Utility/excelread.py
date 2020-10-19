# Program:      Excelread
# Author:       Lex Verheij
# Date:         21 September 2020
# Description:  Reads the Excelsheet and tries to
#               isolate serial numbers & classes

# Documentation for openpyxl:
# https://openpyxl.readthedocs.io/en/stable/

# Imports
from openpyxl import load_workbook
from openpyxl import cell as xlcell

# Global defines
# xlsxPath = "Dummysheet_xlsx.xlsx"
xlsxPath = "80518_Measurement Data_Lot XXX.xlsx"

# load read only workbook without formulas
wb = load_workbook(filename = xlsxPath, data_only = True, read_only = True)

# grab worksheet
ws = wb.active
# ws = wb[xlsxWorksheetName]

# Grab column B and C
range = ws["B1:C500"]   # contains rows with value: (index, serialnr)
index = 0               # first value
serialnr = 1            # second value

# Print the serial numbers in a row with numeric index
for rows in range:
    if type(rows[index]) is xlcell.read_only.EmptyCell: # case that the cell is empty
        continue
    elif (type(rows[index].value) is int): # case that the cell is an integer (ID)
        print("ID", rows[index].value, "has serial number", rows[serialnr].value)
