# subst X: "D:\Technische Informatica\2020-2021\Project GTM\Python"

import socket as SOCKET
import time
from openpyxl import load_workbook
from openpyxl import cell as xlcell

ip = "192.168.0.10"
port = 10000
s = None

s_size = 255                # Can receive up to 255 bytes
s_ipv4 = SOCKET.AF_INET     # Use ipv4 configuration
s_tcp = SOCKET.SOCK_STREAM  # Use TCP configuration

print("Reading Excel sheet...", end="")

xlsxPath = "80518_Measurement Data_Lot XXX.xlsx"
wb = load_workbook(filename = xlsxPath, data_only = True, read_only = True)
ws = wb.active

range = ws["M1:O90"]    # Contains rows with value: (grade, blank, serialnr)
grade = 0               # Index of the first value
serialnr = 2            # Index of the third value
seriallist = []         # List to contain all data (serialnr, grade)

# Print the serial numbers in a row with numeric index
for rows in range:
    if type(rows[grade]) is xlcell.read_only.EmptyCell: # case that the cell is empty
        continue
    elif (type(rows[grade].value) is int): # case that the cell is an integer (ID)
        # print("Cell", rows[serialnr].value, "has grade", rows[grade].value)
        seriallist.append((rows[serialnr].value, rows[grade].value))

print("Done\nConnecting to QR Camera...", end="")

s = SOCKET.socket(s_ipv4, s_tcp)
s.connect((ip, port))
print("Done\nQR Camera connected at ", ip, ":", port, sep="", end=".\n")

firstLine = s.recv(s_size).decode()
# print(">", firstLine)

if "User" not in firstLine:
    # print(">", s.recv(s_size).decode())
    s.recv(s_size)
s.send(b"admin\r\n")
# print(">", s.recv(s_size).decode())
s.recv(s_size)
s.send(b"\r\n")
print(">", s.recv(s_size).decode())
# s.recv(s_size)

waitForInput = True
while True:
    if waitForInput == True:
        input()
    
    waitForInput = True
    
    s.send(b"sw8\r\n")
    s.recv(s_size) # -2
    time.sleep(1)

    s.send(b"gvb002\r\n")
    data = s.recv(s_size)
    sid = data.decode().split("\r\n")[1]
    
    if (len(sid) == 13):
        print("Serial ID =", sid)
    else:
        waitForInput = False
        continue

    for item in seriallist:
        if item[0] == sid:
            print("Grade =", item[1], end="")
            break
    else:
        print("Not Found:", sid, end="")
    
    print("")
