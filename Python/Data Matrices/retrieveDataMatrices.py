import cv2
import urllib.request as u
from openpyxl import load_workbook
from openpyxl import cell as xlcell

xlsxPath = "80518_Measurement Data_Lot XXX.xlsx"
excelData = {}

wb = load_workbook(filename = xlsxPath, data_only = True, read_only = True)
ws = wb.active

# Grab columns M, N and O (col M has grades, col O has serial numbers)
rows = ws["M1:O500"]

# Collect all serial numbers and their corresponding grades
for row in rows:
    if type(row[0]) is not xlcell.read_only.EmptyCell and type(row[0].value) is int:
        grade = row[0].value
        serialnr = row[2].value
        # print("Cell", serialnr, "has grade", grade)
        excelData[serialnr] = grade
wb.close()

for sid in excelData:
    url = "http://datamatrix.kaywa.com/img.php?s=3&d=" + sid.replace(" ", "%20")
    temp = u.urlretrieve(url)[0]
    img = cv2.imread(temp)
    cv2.imwrite(sid + ".png", img)
