# subst X: "D:\Technische Informatica\2020-2021\Project GTM\Python"

"""
Voor het eerste bedrijfsproject in de minor heeft GTM gevraagd of wij een systeem kunnen bouwen die trays met zonnecellen en een excelsheet accepteert en dat omzet naar gesorteerde trays met ieder een bepaalde klasse cellen. De robot zal initieel verbinding maken met de TCP server van de python pc en reageren op elk bericht dat de pc stuurt.
Gedurende het programma zal de robot vanuit de rustpositie wachten tot de pc uit de invoercamera beeld heeft ontvangen en de robot vertelt waar de ongesorteerde cellen liggen. De robot pakt dan een voor een elke ongesorteerde cel op en verplaatst deze boven de qrcamera. De robot zal de pc een bericht sturen dat deze op locatie is.
De pc zal het serienummer laten scannen door de qrcamera en koppelen aan deze cel in de eerste foto. Vervolgens wordt de klasse opgezocht in de excelsheet en stuurt de pc de klasse naar de robot. De robot verplaatst de cel dan naar de uitvoertray met cellen van die klasse. De robot verplaatst zich dan weer naar de rustpositie.
Mocht een uitvoertray vol zitten, laat de robot dit via de pc weten en stopt met verdergaan totdat hij expliciet te horen krijgt dat de uitvoertray is verwijderd. Dit gaat via de hmi. Vervolgens plaatst de robot een lege tray op de geleegde plek en vervolgt hij zijn programma.
Mocht de invoertray leeg zijn, dan vertelt de pc niet welke cel de robot moet pakken, maar dat de tray leeg is en vervangen moet worden. De robot verplaatst dan de lege tray en gaat terug naar de rustpositie. De robot laat vervolgens de pc vertellen waar de volgende cellen liggen en gaat deze een na een af.
"""

import re                               # Regular expressions
import os                               # File path functions, for exists and remove functions
import sys                              # System variables, like error values
import cv2                              # Opencv, for image extraction
import time                             # To measure time intervals
import socket                           # To connect to all devices with the TCP protocol
import numpy as np                      # Numpy is used for opencv image definition and alteration
from ftplib import FTP                  # FTP protocol to download images from the cameras
from queue import SimpleQueue           # Simple queue data type for storing the cell photos
from collections import namedtuple      # Datatype for easy access to device properties
from openpyxl import load_workbook      # Function to import Excel sheets
from openpyxl import cell as xlcell     # Compare and extract Excel cell properties

# Documentation for openpyxl: https://openpyxl.readthedocs.io/en/stable/
# Documentation for opencv:   https://docs.opencv.org/master/

# Path to Excel sheet and a dict for its contents (where (k, v) is (sid, grade))
xlsxPath = "80518_Measurement Data_Lot XXX.xlsx"
excelData = {}

# Socket properties
s_size = 255                # Can receive up to 255 bytes
s_ipv4 = socket.AF_INET     # Use ipv4 configuration
s_tcp = socket.SOCK_STREAM  # Use TCP configuration

# Other variables
timeout = 10                # Seconds
lastRefresh = time.time()   # Timestamp of last refresh
lastPickupInstruction = ""  # Last instruction sent to the robot to pickup a cell/tray
photos = SimpleQueue()      # Store the photos of ungraded cells in a queue

#####################
# HMI message class #
#####################

# HMI Message object for easy read/write access.
class HMImessage:
    """All fields in a HMImessage object retain their values
    // air [0, 1]: turn air pressure on/off [out]
    // light [0, 1]: turn led strip on/off [out]
    // excel [0, 1, 2, 3]: reserved [inout]
    // grade [7-10, 100]: the grade of the cell that has just been classified [inout]
    // outputTray [7-10, 100]: when that output tray needs replacement [inout]
    // trays [0, 1-6]: start signal, the amount of trays stacked on the input pile [in]
    // halt [0, 1]: whether the user asked for a soft stop via the HMI [in]
    // status [0, 1, 2]: either not started, started or finished [inout]
    // conn... [0, 1]: whether the connections with these devices are still live [out]
    """
    def __init__(self):
        self.air, self.light, self.excel, self.grade = 0, 0, 0, 100
        self.outputTray, self.trays, self.halt, self.status = 100, 0, 0, 0
        self.connRobot, self.connQRcam, self.connINcam = 0, 0, 0
    
    # Decodes and stores the received message
    def fromString(self, msg):
        (self.air, self.light, self.excel, self.grade, self.outputTray, self.trays, self.halt,
            self.status, self.connRobot, self.connQRcam, self.connINcam) = [ord(i) for i in msg]
    
    # Concatenate the properties of this object into a string (without encoding)
    def toString(self):
        return (chr(self.air) + chr(self.light) + chr(self.excel) + chr(self.grade) +
            chr(self.outputTray) + chr(self.trays) + chr(self.halt) + chr(self.status) +
                chr(self.connRobot) + chr(self.connQRcam) + chr(self.connINcam))

# Instantiate the class
HMImsg = HMImessage()

#################
# Counter class #
#################

class Counter:
    """Counter object that keeps track of the occupied output spaces"""
    def __init__(self):
        self.outputs = [1, 1, 1, 1]     # Index 0 refers to grade 7, index 3 to grade 10
        self.emptyHeight = 0
        self.inputHeight = 0
    
    # Increments the counter for the output tray of this grade and returns its value after incrementing
    def increment(self, grade):
        self.outputs[grade - 7] += 1
        return self.outputs[grade - 7]
    
    # Resets the counter for the output tray of this grade to 0
    def reset(self, grade):
        self.outputs[grade - 7] = 1

# Instantiate the class
counter = Counter()

#############
# Functions #
#############

def readExcel():
    """Read the Excel sheet and collect its data"""
    wb = load_workbook(filename = xlsxPath, data_only = True, read_only = True)
    ws = wb.active

    # Grab columns M, N and O (col M has grades, col O has serial numbers)
    rows = ws["M1:O500"]

    # Collect all serial numbers and their corresponding grades
    for row in rows:
        if type(row[0]) is not xlcell.read_only.EmptyCell and type(row[0].value) is int:
            grade = row[0].value
            sid = row[2].value
            excelData[sid] = grade
    wb.close()

def retrievePhoto():
    """FTP Retrieve photo function"""
    ftp = FTP(invoerCamera.ip, user="admin", timeout=3)
    with open("image.jpg", "wb") as imageBuffer:
        ftp.retrbinary("RETR image.jpg", imageBuffer.write)
        imageBuffer.close()
    ftp.quit()
    return cv2.imread("image.jpg")

def locateCells(img):
    """Opencv locate cells function"""
    kernel = np.ones((5, 5), np.uint8)
    result = []

    # Cut the image to isolate each cell
    cuts = [0]*12
    cuts[0] = img[165:279, :201]
    cuts[1] = img[165:279, 201:404]
    cuts[2] = img[165:279, 404:603]
    cuts[3] = img[165:279, 603:]
    cuts[4] = img[279:383, :201]
    cuts[5] = img[279:383, 201:404]
    cuts[6] = img[279:383, 404:603]
    cuts[7] = img[279:383, 603:]
    cuts[8] = img[383:497, :201]
    cuts[9] = img[383:497, 201:404]
    cuts[10] = img[383:497, 404:603]
    cuts[11] = img[383:497, 603:]

    # Decide for every spot if it is occupied by a cell
    for i in range(12):
        org = cuts[i]
        img = cv2.cvtColor(org, cv2.COLOR_BGR2GRAY)                         # Grayscale
        img = cv2.threshold(img, 64, 255, cv2.THRESH_BINARY)[1]             # BW
        img = cv2.morphologyEx(img, cv2.MORPH_OPEN, kernel, iterations=2)   # Close holes
        if cv2.countNonZero(img) < 15000:   # If there are less than 15000 pixels in this cut
            photos.put(org)                 # Temporarily store this image, until we know its grade
            result.append(i + 1)            # Cell 0 has position 1, cell 11 has position 12, etc.
    return result

def SavePhoto(sid):
    """Save the photo of this cell with its serial number as file name"""
    cv2.imwrite("Photos\\" + sid + ".png", photos.get(block=False))

def connect(device):
    """TCP Connect function"""
    device.socket.connect((device.ip, device.port))
    print(device.name, " connected at ", device.ip, ":", device.port, sep="")

def login(device):
    """TCP Login function"""
    if "User" not in device.socket.recv(s_size).decode():   # Welcome to In-Sight(tm)  7200C Session 0
        device.socket.recv(s_size)                          # User:
    device.socket.send(b"admin\r\n")
    device.socket.recv(s_size)                              # Password:
    device.socket.send(b"\r\n")
    device.socket.recv(s_size)                              # User Logged In
    print(device.name, "login success")

def disconnect(device):
    """TCP Disconnect function"""
    device.socket.close()

def handleError(device):
    """TCP Handle Errors function updates the connection info to display on the HMI and disconnects all devices"""
    if device is not None: print("TCP Error for device", device.name, sys.exc_info()[0])
    if device != robot: send(robot, "disconnect"); disconnect(robot)
    if device != qrCamera: disconnect(qrCamera)
    if device != invoerCamera: disconnect(invoerCamera)
    if device != plc:
        if device == robot: HMImsg.connRobot = 0
        elif device == qrCamera: HMImsg.connQRcam = 0
        elif device == invoerCamera: HMImsg.connINcam = 0
        if device is not None: send(plc, HMImsg.toString())
        disconnect(plc)
    exit(-1)

def send(device, msg, verbal = True):
    """TCP Send function"""
    try:
        if verbal: print("PC > ", device.name, ":\t", msg.encode(), sep="")
        device.socket.send(msg.encode())
    except:
        handleError(device)

def receive(device, index = None, verbal = True):
    """TCP Receive function. If index is provided, the received data will be split"""
    if device == plc:
        flush(plc)
    try:
        data = device.socket.recv(s_size)
        if verbal: print(device.name, " > PC:\t", data, sep="")
        data = data.decode()
        if index is not None:
            data = data.split("\r\n")[index]
        return data
    except:
        handleError(device)

def flush(device):
    """TCP flush function that empties its receive buffer to ensure actual data (e.g. HMI keeps sending data)"""
    try:
        while len(device.socket.recv(s_size)) == s_size:
            pass
    except:
        handleError(device)

def refresh(excelloop):
    """TCP Refresh function"""
    global lastRefresh
    while True:
        # While within refresh timeout
        while time.time() - lastRefresh < timeout:
            HMImsg.fromString(receive(plc, verbal = False))

            # Return when the program is running, not halted, without full output trays and without excel errors
            if HMImsg.status and not HMImsg.halt and HMImsg.outputTray == 100 and not HMImsg.excel:
                return
            
            # If the user chooses to terminate the program due to an Excel error
            if HMImsg.excel == 4:
                robotReturnCell()

            # Print some debug messages while waiting for the user to acknowledge via the HMI
            if HMImsg.status != 1: print("Waiting for start acknowledge by user via HMI...")
            if HMImsg.halt != 0: print("Program halted. Waiting for user acknowledge via HMI...")
            if HMImsg.outputTray != 100: print("Waiting for user to acknowledge output tray removal via HMI...")
            if HMImsg.excel == 1: print("Waiting for user to acknowledge new excel sheet upload via HMI...")
            if HMImsg.excel == 2: print("Waiting for user response to serial number not being in excel sheet via HMI...")
            if HMImsg.excel == 3: print("Waiting for user response to unreadable data matrix via HMI...")

            time.sleep(0.5)

        # Refresh all other connections
        print("Refreshing devices")
        lastRefresh = time.time()
        send(robot, "refresh", verbal = False)
        receive(robot, verbal = False)
        send(qrCamera, "refresh\r\n", verbal = False)
        receive(qrCamera, verbal = False)
        send(invoerCamera, "refresh\r\n", verbal = False)
        receive(invoerCamera, verbal = False)

def robotReturnCell():
    """This function instructs the robot to return the cell to its last location and stops the program"""
    send(robot, lastPickupInstruction)
    receive(robot)                  # Robot has reached cell location

    HMImsg.fromString(receive(plc)) # Don't overwrite with cached data. No refresh, since halt (e.g.) corrupts robot movement
    HMImsg.air = 0
    send(plc, HMImsg.toString())    # Instruct the PLC to enable the vacuum generator
    time.sleep(0.1)                 # Give the PLC time to process

    send(robot, "neutral")          # Instruct the robot return to its neutral position
    receive(robot)                  # Robot has reached its destination

    handleError(None)

def robotMsg(instruction, heightGrade, celIndex):
    """Robot Message function for easy message creation using convention 'X-YY-ZZ'"""
    global lastPickupInstruction
    heightGrade = str(heightGrade) if heightGrade > 9 else "0" + str(heightGrade)
    celIndex = str(celIndex) if celIndex > 9 else "0" + str(celIndex)
    res = instruction + "-" + heightGrade + "-" + celIndex
    if res[0] == "I": lastPickupInstruction = res
    return res

def newSocket():
    """Function to create a new socket"""
    s = socket.socket(s_ipv4, s_tcp)
    s.settimeout(timeout)
    return s

#####################
# Device properties #
#####################

# Device object to enable named fields: camera.socket instead of camera[2]
Device = namedtuple("Device", "name, ip, port, socket")

# Properties of the connecting devices (Name, IP, Port, Client socket)
# The PLC and cameras always immediately respond, since their socket handling is non-blocking. The robot will not
plc = Device("PLC", "192.168.0.2", 2000, newSocket())  # Connect to the HMI via the PLC
robot = Device("Robot", "192.168.0.1", 10000, newSocket())
qrCamera = Device("QRcam", "192.168.0.10", 10000, newSocket())
invoerCamera = Device("INcam", "192.168.0.11", 10000, newSocket())

###################
# Connect devices #
###################

# Setup a client socket for the PLC and connect
print("Connecting to PLC...", end="\r", flush=True)
connect(plc)

# Setup a client socket for the Robot and connect
print("Connecting to Robot...", end="\r", flush=True)
connect(robot)
HMImsg.connRobot = 1
send(plc, HMImsg.toString(), verbal = False)

# Setup a client socket for the QR Camera, connect and login
print("Connecting to QR Camera...", end="\r", flush=True)
connect(qrCamera)
login(qrCamera)
HMImsg.connQRcam = 1
send(plc, HMImsg.toString(), verbal = False)

# Setup a client socket for the Invoer Camera, connect and login
print("Connecting to Invoer Camera...", end="\r", flush=True)
connect(invoerCamera)
login(invoerCamera)
HMImsg.connINcam = 1
send(plc, HMImsg.toString(), verbal = False)

####################################################
# Read Excel sheet and communicate errors with HMI #
####################################################

# Since refresh also blocks program flow if status is 0, temporarily change it to 8
HMImsg.status = 8
send(plc, HMImsg.toString(), verbal = False)

# Read Excel sheet and put the data in excelData
while True:
    try:
        print("Attempting to read Excelsheet...", end="\r")
        readExcel()
        print("Excel sheet has been read successfully")
        break
    except FileNotFoundError:
        HMImsg.excel = 1
        send(plc, HMImsg.toString())    # Show HMI warning and wait for response
        time.sleep(0.5)                 # Give the PLC time to process
        refresh()

# Reset the mocked status code
HMImsg.status = 0
send(plc, HMImsg.toString(), verbal = False)
time.sleep(1)                           # Give the PLC time to process

################
# Main program #
################

# Connections are established, so we wait for the user to start the program on the HMI
while HMImsg.status != 1:
    refresh()

# At this point, we received the amount of input trays from the HMI
HMImsg.fromString(receive(plc))
time.sleep(0.1)                         # Give the PLC time to process
counter.inputHeight = HMImsg.trays      # Start the size of the inputHeight at HMImsg.trays
send(robot, "neutral")                  # We then Instruct the robot to move to its neutral position
receive(robot)                          # Robot has reached its neutral position

# Repeat this loop while there are input trays left
while counter.inputHeight:
    refresh()                           # Revalidate all tcp connections
    send(invoerCamera, "sw8\r\n")       # Trigger a photo
    receive(invoerCamera, 0)            # Acknowledge
    locs = locateCells(retrievePhoto()) # Use FTP to retrieve the photo and locate all cells

    # Instruct the robot to go to cell i in the input tray "I" and also provide the height
    for i in locs:
        refresh()
        send(robot, robotMsg("I", counter.inputHeight, i))
        receive(robot)                  # Robot has reached cell location

        HMImsg.fromString(receive(plc)) # Don't overwrite with cached data. No refresh, since halt (e.g.) corrupts robot movement
        HMImsg.air = 1
        send(plc, HMImsg.toString())    # Instruct the PLC to enable the vacuum generator
        time.sleep(0.1)                 # Give the PLC time to process

        refresh()
        send(robot, "scan")             # Instruct the robot to move the cell to the QR camera for a scan
        receive(robot)                  # Robot has reached its destination

        HMImsg.fromString(receive(plc)) # Don't overwrite with cached data. No refresh, since halt (e.g.) corrupts robot movement
        HMImsg.light = 1
        send(plc, HMImsg.toString())    # Instruct the PLC to enable the lighting for the QR camera
        time.sleep(0.1)                 # Give the PLC time to process

        # Read the serial number (data matrix), retry ten times, if needed
        for _ in range(10):
            refresh()
            send(qrCamera, "sw8\r\n")       # Trigger a photo
            receive(qrCamera, 0)            # Acknowledge
            time.sleep(0.3)                 # Give the camera time to process the data matrix
            send(qrCamera, "gvb002\r\n")
            sid = receive(qrCamera, 1)      # Get the serial number from the camera

            # If this sid matches the format of "xxxxx xxxx xx" where all x are digits, get the grade
            if re.search("\d{5} \d{4} \d{2}", sid):
                try:
                    grade = excelData[sid]
                    SavePhoto(sid)
                    break
                # If the sid is not present in the Excel sheet
                except KeyError:
                    print("Excel sheet does not contain <", sid, ">", sep="")
                    refresh()
                    HMImsg.excel = 2
                    HMImsg.light = 0
                    send(plc, HMImsg.toString())    # Instruct the PLC to disable the lights and to show the error message
                    time.sleep(0.1)                 # Give the PLC time to process

                    refresh()                       # Wait for user to accept message on hmi
                    robotReturnCell()               # Return

        # If this else clause is executed, it means that the data matrix is not recognized after 10 retries
        else:
            print("Data matrix cannot be recognized")
            SavePhoto(time.strftime("unreadable cell on %Y-%m-%d %X GMT", time.gmtime()))

            refresh()
            HMImsg.excel = 3
            HMImsg.light = 0
            send(plc, HMImsg.toString())    # Instruct the PLC to disable the lights and to show the error message
            time.sleep(0.1)                 # Give the PLC time to process

            refresh()                       # Wait for user choice (if not quit, then move cell to spare box)
            send(robot, "box")              # Instruct the robot to place the cell there
            receive(robot)                  # Robot has reached its destination

            HMImsg.fromString(receive(plc)) # Don't overwrite with cached data. No refresh, since halt (e.g.) corrupts robot movement
            HMImsg.air = 0
            send(plc, HMImsg.toString())    # Instruct the PLC to disable the vacuum generator
            time.sleep(0.1)                 # Give the PLC time to process

            refresh()
            send(robot, "neutral")          # Instruct the robot return to its neutral position
            receive(robot)                  # Robot has reached its destination
            continue
        
        refresh()
        HMImsg.grade = grade            # Store this grade in the HMImsg object
        HMImsg.light = 0                # Also, the light can be turned off again
        send(plc, HMImsg.toString())    # Instruct the PLC to enable the lighting for the QR camera
        time.sleep(0.1)                 # Give the PLC time to process

        refresh()
        send(robot, "neutral")          # Instruct the robot return to its neutral position
        receive(robot)                  # Robot has reached its destination
        
        refresh()
        cell = counter.increment(grade) - 1     # Increment output tray cell counter and store free spot
        send(robot, robotMsg("Q", grade, cell)) # Instruct the robot to place the cell there
        receive(robot)                  # Robot has reached its destination

        HMImsg.fromString(receive(plc)) # Don't overwrite with cached data. No refresh, since halt (e.g.) corrupts robot movement
        HMImsg.air = 0
        send(plc, HMImsg.toString())    # Instruct the PLC to disable the vacuum generator
        time.sleep(0.1)                 # Give the PLC time to process

        refresh()
        send(robot, "neutral")          # Instruct the robot return to its neutral position
        receive(robot)                  # Robot has reached its neutral position

        refresh()
        HMImsg.grade = 100
        send(plc, HMImsg.toString())    # Instruct the PLC to remove the grade value from the HMI
        time.sleep(0.1)                 # Give the PLC time to process

        # If output tray is full, the 12th cell has just been placed in that output tray
        if cell == 12:
            refresh()
            counter.reset(grade)        # The user will physically remove the tray, reset counter to 0
            HMImsg.outputTray = grade
            send(plc, HMImsg.toString())    # Instruct the HMI to execute the tray replacement procedure
            time.sleep(0.1)                 # Give the PLC time to process

    # If input tray has no cells left, move the tray to the empty tray stack
    refresh()
    send(robot, robotMsg("I", counter.inputHeight, 0))
    counter.inputHeight -= 1        # Robot picks up the input tray
    receive(robot)                  # Robot has reached its destination

    HMImsg.fromString(receive(plc)) # Don't overwrite with cached data. No refresh, since halt (e.g.) corrupts robot movement
    HMImsg.air = 1
    send(plc, HMImsg.toString())    # Instruct the PLC to enable the vacuum generator
    time.sleep(0.1)                 # Give the PLC time to process

    refresh()
    counter.emptyHeight += 1        # Robot places tray on the stack at the new height
    send(robot, robotMsg("S", counter.emptyHeight, 0))
    receive(robot)                  # Robot has reached its destination

    HMImsg.fromString(receive(plc)) # Don't overwrite with cached data. No refresh, since halt (e.g.) corrupts robot movement
    HMImsg.air = 0
    send(plc, HMImsg.toString())    # Instruct the PLC to disable the vacuum generator
    time.sleep(0.1)                 # Give the PLC time to process

    refresh()
    send(robot, "neutral")          # Instruct the robot return to its neutral position
    receive(robot)                  # Robot has reached its neutral position

# Program finished
HMImsg.status = 4
send(plc, HMImsg.toString())
time.sleep(1)                       # Give the PLC time to process

######################
# Disconnect devices #
######################

send(robot, "disconnect")
disconnect(robot)
disconnect(plc)
disconnect(qrCamera)
disconnect(invoerCamera)
